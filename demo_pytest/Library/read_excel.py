from openpyxl import load_workbook
from Library.configure import Configuration

class ReadExcel:

    def locators_details(self):
        path = Configuration.LOCATORS_PATH
        wb = load_workbook(path)
        ws = wb[Configuration.SHEETNAME_LOCATORS]
        data = {}
        for row in ws.iter_rows(min_row=2):
            data[row[0].value] = (row[1].value, row[2].value)
        return data


    def login_details(self):
        path = Configuration.LOCATORS_PATH
        wb = load_workbook(path)
        ws = wb[Configuration.SHEETNAME_DATA]
        l_ = []
        for row in ws.iter_rows(min_row=2):
            l_.append((row[0].value, row[1].value))
        return l_

r = ReadExcel()
data = r.login_details()
print(data)
data1 = r.locators_details()
print(data1)


